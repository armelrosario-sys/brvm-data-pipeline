#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""P5 — Generateur du tableau de bord Excel.
Regenere entierement a chaque run depuis le moteur (scoring.py) — jamais
edite a la main. 4 onglets : Synthese, Fiches titres, Alertes, Registre.
"""
import sys
from datetime import date, datetime
from pathlib import Path
import sqlite3

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "moteur"))
from scoring import evaluer_titre, rapport_fraicheur, charger_seuils, DB

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Calibri"
C_ENTETE = "1F2937"
C_ELIGIBLE = "D1FAE5"
C_EXCLU = "FEE2E2"
C_VIGILANCE = "FEF3C7"
C_BLANC = "FFFFFF"

BORDURE = Border(*(Side(style="thin", color="D1D5DB"),) * 4)


def style_entete(cell):
    cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", start_color=C_ENTETE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDURE


def style_cellule(cell, gras=False, couleur_fond=None, centre=False):
    cell.font = Font(name=FONT, bold=gras, size=10)
    if couleur_fond:
        cell.fill = PatternFill("solid", start_color=couleur_fond)
    cell.alignment = Alignment(horizontal="center" if centre else "left", vertical="center")
    cell.border = BORDURE


def largeurs(ws, largeurs_dict):
    for col, w in largeurs_dict.items():
        ws.column_dimensions[col].width = w


def onglet_synthese(wb, resultats, seuils_version):
    ws = wb.active
    ws.title = "Synthese"
    ws["A1"] = "BRVM — Watchlist GARP-Quality"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ws["A2"] = f"Genere le {datetime.now():%d/%m/%Y %H:%M} · seuils v{seuils_version} · NE SERT A AUCUNE DECISION D'INVESTISSEMENT SEULE"
    ws["A2"].font = Font(name=FONT, italic=True, size=9, color="B91C1C")
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")

    entetes = ["Rang", "Titre", "Secteur", "Statut", "Score", "Rentabilite",
               "Solidite", "Valorisation", "Sizing", "Alertes principales"]
    r0 = 4
    for j, h in enumerate(entetes, 1):
        style_entete(ws.cell(row=r0, column=j, value=h))

    elig = [r for r in resultats if r["statut_gate"] == "ELIGIBLE"]
    excl = [r for r in resultats if r["statut_gate"] == "EXCLU"]
    elig.sort(key=lambda r: -(r["score_composite"] or 0))

    row = r0 + 1
    for i, r in enumerate(elig, 1):
        alertes_maj = [a for a in r["alertes"] if any(
            m in a for m in ("TURNAROUND", "PERIMEE", "payout", "recul", "AVIS", "ROE"))]
        vals = [i, r["ticker"], r["secteur"], "ELIGIBLE",
                round(r["score_composite"], 1) if r["score_composite"] else "—",
                round(r["score_rentabilite"], 1) if r["score_rentabilite"] else "—",
                round(r["score_solidite"], 1) if r["score_solidite"] else "—",
                round(r["score_valorisation"], 1) if r["score_valorisation"] else "—",
                r["sizing"]["recommandation"], "; ".join(alertes_maj[:2])]
        for j, v in enumerate(vals, 1):
            fond = C_ELIGIBLE if j <= 4 else (C_VIGILANCE if alertes_maj and j == 10 else None)
            style_cellule(ws.cell(row=row, column=j, value=v), centre=(j <= 5 or j == 9))
        row += 1

    for r in excl:
        vals = [None, r["ticker"], r["secteur"], "EXCLU", "—", "—", "—", "—", "—",
                "; ".join(r["motifs_exclusion"][:2])]
        for j, v in enumerate(vals, 1):
            style_cellule(ws.cell(row=row, column=j, value=v), centre=(j <= 5 or j == 9),
                          couleur_fond=C_EXCLU if j <= 4 else None)
        row += 1

    largeurs(ws, {"A": 6, "B": 9, "C": 22, "D": 10, "E": 8, "F": 11,
                  "G": 10, "H": 12, "I": 12, "J": 60})
    ws.freeze_panes = "A5"
    return elig, excl


def onglet_fiches(wb, resultats):
    ws = wb.create_sheet("Fiches titres")
    entetes = ["Titre", "Secteur", "Statut", "Score", "Alertes / motifs (detail complet)"]
    for j, h in enumerate(entetes, 1):
        style_entete(ws.cell(row=1, column=j, value=h))
    row = 2
    for r in sorted(resultats, key=lambda x: x["ticker"]):
        detail = r["motifs_exclusion"] if r["statut_gate"] == "EXCLU" else r["alertes"]
        vals = [r["ticker"], r["secteur"], r["statut_gate"],
                round(r["score_composite"], 1) if r.get("score_composite") else "—",
                "\n".join(detail) if detail else "(aucune)"]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=j, value=v)
            style_cellule(c, centre=(j <= 4))
            if j == 5:
                c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(15, 14 * len(detail))
        row += 1
    largeurs(ws, {"A": 9, "B": 22, "C": 10, "D": 8, "E": 90})
    ws.freeze_panes = "A2"


def onglet_alertes(wb, resultats):
    ws = wb.create_sheet("Alertes")
    entetes = ["Titre", "Type de signal", "Detail"]
    for j, h in enumerate(entetes, 1):
        style_entete(ws.cell(row=1, column=j, value=h))
    row = 2
    categories = [("TURNAROUND", "Turnaround"), ("PERIMEE", "Donnee perimee"),
                  ("payout", "Payout"), ("recul", "Recul resultat/CP"),
                  ("AVIS", "Avis reglementaire"), ("PROFIT_WARNING", "Profit warning")]
    for r in resultats:
        for a in r["alertes"]:
            for cle, label in categories:
                if cle in a:
                    for j, v in enumerate([r["ticker"], label, a], 1):
                        style_cellule(ws.cell(row=row, column=j, value=v), centre=(j <= 2))
                    row += 1
                    break
        if r["statut_gate"] == "EXCLU":
            for m in r["motifs_exclusion"]:
                for j, v in enumerate([r["ticker"], "EXCLUSION", m], 1):
                    style_cellule(ws.cell(row=row, column=j, value=v),
                                  centre=(j <= 2), couleur_fond=C_EXCLU if j == 2 else None)
                row += 1
    largeurs(ws, {"A": 9, "B": 20, "C": 90})
    ws.freeze_panes = "A2"


def onglet_registre(wb, seuils, fraicheur):
    ws = wb.create_sheet("Registre")
    ws["A1"] = "Registre des seuils et etat de fraicheur"
    ws["A1"].font = Font(name=FONT, bold=True, size=13)
    ws["A2"] = f"Seuils version {seuils.get('version', '?')} — {seuils.get('date', '?')}"
    ws["A2"].font = Font(name=FONT, italic=True, size=10)

    r = 4
    for h in ["Categorie", "Cle", "Valeur"]:
        style_entete(ws.cell(row=r, column=["Categorie", "Cle", "Valeur"].index(h) + 1, value=h))
    r += 1
    for cat, bloc in seuils.items():
        if not isinstance(bloc, dict):
            continue
        for k, v in bloc.items():
            if isinstance(v, dict):
                v = str(v)
            style_cellule(ws.cell(row=r, column=1, value=cat))
            style_cellule(ws.cell(row=r, column=2, value=k))
            style_cellule(ws.cell(row=r, column=3, value=str(v)), centre=True)
            r += 1

    r += 2
    ws.cell(row=r, column=1, value="Etat de fraicheur (R7)").font = Font(name=FONT, bold=True, size=12)
    r += 1
    for h in ["Titre", "Dernier mois", "Jours", "Statut"]:
        style_entete(ws.cell(row=r, column=["Titre", "Dernier mois", "Jours", "Statut"].index(h) + 1, value=h))
    r += 1
    for t, fm, j in fraicheur["perimes"]:
        for col, v in enumerate([t, fm, j, "PERIME"], 1):
            style_cellule(ws.cell(row=r, column=col, value=v), centre=True,
                          couleur_fond=C_VIGILANCE if col == 4 else None)
        r += 1
    for t, fm, j in fraicheur["frais"][:5]:
        for col, v in enumerate([t, fm, j, "OK"], 1):
            style_cellule(ws.cell(row=r, column=col, value=v), centre=True)
        r += 1
    largeurs(ws, {"A": 26, "B": 14, "C": 10, "D": 12})


def main():
    conn = sqlite3.connect(DB)
    tickers = [row[0] for row in conn.execute(
        "SELECT ticker FROM societes WHERE ticker NOT LIKE 'TEST_%'").fetchall()]
    conn.close()
    resultats = [evaluer_titre(t) for t in tickers]
    seuils = charger_seuils()
    fraicheur = rapport_fraicheur()

    wb = Workbook()
    onglet_synthese(wb, resultats, seuils.get("version", "?"))
    onglet_fiches(wb, resultats)
    onglet_alertes(wb, resultats)
    onglet_registre(wb, seuils, fraicheur)

    sortie = Path(__file__).resolve().parent.parent / "dashboard_brvm.xlsx"
    wb.save(sortie)
    print(f"Dashboard genere : {sortie} ({len(resultats)} titres)")


if __name__ == "__main__":
    main()
